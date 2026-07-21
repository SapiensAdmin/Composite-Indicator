"""
build_excel.py — assemble the 4-sheet workbook `data/liquidity_composite.xlsx`.

Sheets:
  A) data_long        — tidy fact table (single source of truth for raw data)
  B) series_registry  — the composite control panel (hand-edited)
  C) composite        — computed output (filled by compute_composite.py)
  D) top10_membership — audit trail of which schemes were top-10-by-AUM each month

Run standalone to create/refresh the workbook from whatever is currently on disk
(empty data is fine — this is build-order step 2). The writer PRESERVES an existing
registry sheet, so re-running never clobbers hand-edited weights.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import common as C


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
REGISTRY_FILL = PatternFill("solid", fgColor="0E7490")


def _style_sheet(ws, highlight: bool = False) -> None:
    """Bold coloured header row, frozen panes, sensible column widths, autofilter."""
    if ws.max_row == 0:
        return
    fill = REGISTRY_FILL if highlight else HEADER_FILL
    for cell in ws[1]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        values = [str(c.value) if c.value is not None else "" for c in col_cells]
        width = min(max((len(v) for v in values), default=10) + 3, 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(frames: dict[str, pd.DataFrame]) -> None:
    """Write all four sheets to WORKBOOK_PATH with light formatting."""
    C.DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Deterministic ordering makes diffs (and the audit trail) readable.
    data_long = frames[C.SHEET_DATA_LONG].copy()
    if not data_long.empty:
        data_long = data_long.sort_values(
            ["as_of_date", "category", "scheme_name", "metric"]
        ).reset_index(drop=True)

    top10 = frames[C.SHEET_TOP10].copy()
    if not top10.empty:
        top10 = top10.sort_values(["as_of_date", "category", "rank"]).reset_index(drop=True)

    ordered = {
        C.SHEET_DATA_LONG: data_long,
        C.SHEET_REGISTRY: frames[C.SHEET_REGISTRY],
        C.SHEET_COMPOSITE: frames[C.SHEET_COMPOSITE],
        C.SHEET_TOP10: top10,
    }

    with pd.ExcelWriter(WORKBOOK_PATH_STR, engine="openpyxl") as writer:
        for sheet, df in ordered.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
        wb = writer.book
        for sheet in ordered:
            _style_sheet(wb[sheet], highlight=(sheet == C.SHEET_REGISTRY))


WORKBOOK_PATH_STR = str(C.WORKBOOK_PATH)


def ensure_workbook() -> None:
    """Create the workbook with the seeded registry if it doesn't exist yet."""
    frames = C.load_frames()
    write_workbook(frames)


if __name__ == "__main__":
    ensure_workbook()
    frames = C.load_frames()
    print(f"Workbook ready at {C.WORKBOOK_PATH}")
    for name, df in frames.items():
        print(f"  {name:18s}: {len(df):5d} rows, cols={list(df.columns)}")
